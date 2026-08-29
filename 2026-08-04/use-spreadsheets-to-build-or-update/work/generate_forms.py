import argparse
import os
from collections import Counter

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas


SOURCE_W = 612.0
SOURCE_H = 1008.0
A4_W, A4_H = A4
SCALE = min(A4_W / SOURCE_W, A4_H / SOURCE_H)
OFFSET_X = (A4_W - SOURCE_W * SCALE) / 2
OFFSET_Y = (A4_H - SOURCE_H * SCALE) / 2


def clean(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def hh_id(value):
    v = clean(value)
    if v.startswith("#"):
        v = v[1:]
    if v.endswith(".0") and v[:-2].isdigit():
        v = v[:-2]
    return v


def roster_records(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["FR 05312026"]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    ix = {name: idx for idx, name in enumerate(headers)}
    wanted = [
        "HH_ID", "LAST_NAME", "FIRST_NAME", "MIDDLE_NAME", "EXT_NAME",
        "REGION", "PROVINCE", "MUNICIPALITY", "BARANGAY", "CLIENT_STATUS",
        "GRANTEE",
    ]
    missing = [name for name in wanted if name not in ix]
    if missing:
        raise ValueError(f"Required roster columns missing: {missing}")

    records = []
    for row in rows:
        if clean(row[ix["GRANTEE"]]).upper() != "GRANTEE":
            continue
        records.append({name: clean(row[ix[name]]) for name in wanted})
    wb.close()
    records.sort(key=lambda x: (x["BARANGAY"], x["LAST_NAME"], x["FIRST_NAME"], hh_id(x["HH_ID"])))
    return records


def page_xy(x, y):
    return OFFSET_X + x * SCALE, OFFSET_Y + y * SCALE


def draw_fit(c, text, x, y, max_width, source_font=7.8, font="Helvetica-Bold"):
    text = clean(text)
    if not text:
        return
    font_size = source_font * SCALE
    width = max_width * SCALE
    while font_size > 4.4 and stringWidth(text, font, font_size) > width:
        font_size -= 0.2
    c.setFont(font, font_size)
    c.drawString(*page_xy(x, y), text)


def full_name(rec):
    parts = [rec["LAST_NAME"] + (f", {rec['FIRST_NAME']}" if rec["FIRST_NAME"] else "")]
    if rec["MIDDLE_NAME"]:
        parts.append(rec["MIDDLE_NAME"])
    if rec["EXT_NAME"]:
        parts.append(rec["EXT_NAME"])
    return " ".join(p for p in parts if p)


def add_page_one(c, background, rec):
    c.drawImage(background, OFFSET_X, OFFSET_Y, SOURCE_W * SCALE, SOURCE_H * SCALE, preserveAspectRatio=True)

    draw_fit(c, rec["LAST_NAME"], 95, 868.5, 145)
    draw_fit(c, rec["FIRST_NAME"], 246, 868.5, 185)
    draw_fit(c, rec["MIDDLE_NAME"], 438, 868.5, 91)
    draw_fit(c, rec["EXT_NAME"], 535, 868.5, 60)

    draw_fit(c, hh_id(rec["HH_ID"]), 120, 850.0, 172, 8.2)
    # The roster has no house number or street/purok/sitio fields, so those remain blank.
    draw_fit(c, rec["BARANGAY"], 90, 830.2, 150)
    draw_fit(c, rec["MUNICIPALITY"], 290, 830.2, 137)
    draw_fit(c, rec["PROVINCE"], 440, 830.2, 91)
    draw_fit(c, rec["REGION"], 555, 830.2, 40)

    name = full_name(rec)
    hid = hh_id(rec["HH_ID"])
    draw_fit(c, name, 110, 239.5, 70, 6.6)
    draw_fit(c, hid, 206, 239.5, 62, 6.6)
    draw_fit(c, name, 390, 239.5, 72, 6.6)
    draw_fit(c, hid, 500, 239.5, 55, 6.6)
    c.showPage()


def add_page_two(c, background):
    c.drawImage(background, OFFSET_X, OFFSET_Y, SOURCE_W * SCALE, SOURCE_H * SCALE, preserveAspectRatio=True)
    c.showPage()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--roster", required=True)
    parser.add_argument("--page1", required=True)
    parser.add_argument("--page2", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--limit", type=int)
    args = parser.parse_args()

    records = roster_records(args.roster)
    if args.limit is not None:
        records = records[: args.limit]
    if not records:
        raise ValueError("No grantee records found")

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    c = canvas.Canvas(args.output, pagesize=A4, pageCompression=1)
    c.setTitle("BUS Form 5 - San Fabian Family Roster as of 31 May 2026")
    c.setAuthor("DSWD MOO San Fabian")
    c.setSubject("A4-printable BUS Form 5 forms populated from the family roster")
    for rec in records:
        add_page_one(c, args.page1, rec)
        add_page_two(c, args.page2)
    c.save()

    statuses = Counter(r["CLIENT_STATUS"] for r in records)
    print(f"households={len(records)} pages={len(records) * 2}")
    for status, count in sorted(statuses.items()):
        print(f"{count}\t{status}")


if __name__ == "__main__":
    main()
