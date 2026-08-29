from openpyxl import load_workbook
p=r"C:\Users\WINDOWS\Downloads\MOO San Fabian - Family Roster as of 31 May 2026.xlsx"
wb=load_workbook(p,read_only=True,data_only=True)
ws=wb["FR 05312026"]
print(list(next(ws.iter_rows(min_row=1,max_row=1,values_only=True))))
