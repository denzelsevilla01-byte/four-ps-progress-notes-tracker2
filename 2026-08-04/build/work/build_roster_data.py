from openpyxl import load_workbook
from collections import Counter
from datetime import date, datetime
from pathlib import Path
import sys
import json

source = Path(r"C:\Users\WINDOWS\Downloads\MOO San Fabian - Family Roster as of 31 May 2026.xlsx")
target = Path(next((a.split("=",1)[1] for a in sys.argv if a.startswith("--target=")), r"C:\Users\WINDOWS\Documents\Codex\2026-08-04\build\outputs\san-fabian-roster-data.js"))
wb = load_workbook(source, read_only=True, data_only=True)
ws = wb["FR 05312026"]
rows = ws.iter_rows(values_only=True)
headers = [str(x or "") for x in next(rows)]
ix = {h:i for i,h in enumerate(headers)}

def val(row, key):
    x = row[ix[key]]
    if x is None: return ""
    if isinstance(x, (date, datetime)): return x.isoformat()[:10]
    return str(x).strip()

households = {}
for row in rows:
    hh = val(row, "HH_ID")
    if not hh: continue
    if hh.endswith(".0"): hh = hh[:-2]
    rec = households.get(hh)
    if rec is None:
        rec = {"id":hh,"set":val(row,"HH_SET"),"status":val(row,"CLIENT_STATUS"),"cluster":val(row,"CLUSTER"),"barangay":val(row,"BARANGAY"),"grantee":"","members":[],"flags":set()}
        households[hh] = rec
    member = [val(row,k) for k in ["FULL_NAME","RELATION_TO_HH_HEAD","AGE","SEX","MEMBER_STATUS","CHILD_BENE","ATTEND_SCHOOL","SCHOOL_NAME","GRADE_LEVEL","HEALTH_MONITORED","HEALTH_FACILITY","DISABILITY_TYPES","PREGNANCY_STATUS","SOLO_PARENT","IP_AFFILIATION","PCN","BIRTHDAY","CIVIL_STATUS","EDUC_MONITORED","LRN","HEALTH_FACILITY_STATUS"]]
    rec["members"].append(member)
    if val(row,"GRANTEE").upper() == "GRANTEE": rec["grantee"] = val(row,"FULL_NAME")
    status = rec["status"].lower()
    if "on-hold" in status or "on hold" in status: rec["flags"].add("On hold")
    if "unlocated" in status or "moved out" in status: rec["flags"].add("Locate / validate")
    if "processing" in status: rec["flags"].add("For processing")
    if val(row,"DISABILITY_TYPES") and val(row,"DISABILITY_TYPES").upper() not in {"NO","NONE","N/A"}: rec["flags"].add("PWD member")
    if val(row,"PREGNANCY_STATUS") and val(row,"PREGNANCY_STATUS").upper() not in {"NO","NOT PREGNANT","N/A"}: rec["flags"].add("Pregnancy monitoring")
    if val(row,"SOLO_PARENT").upper() in {"YES","Y"}: rec["flags"].add("Solo parent")
    if val(row,"CHILD_BENE").upper() not in {"","NO"} and val(row,"ATTEND_SCHOOL").upper() in {"NO","NOT ATTENDING"}: rec["flags"].add("Education follow-up")

status_counts = Counter()
barangay_counts = Counter()
member_status_counts = Counter()
out = []
for h in households.values():
    status_counts[h["status"]] += 1
    barangay_counts[h["barangay"]] += 1
    for m in h["members"]: member_status_counts[m[4]] += 1
    out.append([h["id"],h["set"],h["status"],h["cluster"],h["barangay"],h["grantee"],sorted(h["flags"]),h["members"]])
out.sort(key=lambda x:(x[4],x[5],x[0]))
payload={"asOf":"31 May 2026","households":out,"meta":{"households":len(out),"members":sum(len(h[7]) for h in out),"statuses":status_counts,"barangays":barangay_counts,"memberStatuses":member_status_counts}}
content = "window.ROSTER_DATA="+json.dumps(payload,ensure_ascii=False,separators=(",",":"))+";"
if "--stdout" in sys.argv:
    sys.stdout.write(content)
else:
    target.write_text(content,encoding="utf-8")
    print(json.dumps({"households":payload["meta"]["households"],"members":payload["meta"]["members"],"statuses":len(status_counts),"barangays":len(barangay_counts),"bytes":target.stat().st_size},indent=2))
